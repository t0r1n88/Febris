import tkinter
import sys
import warnings
import os
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
import time
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
pd.options.mode.chained_assignment = None  # default='warn'
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


"""
классы для исключений
"""

class WrongNumberColumn(Exception):
    """
    Класс для исключения в случае неправильного количества колонок в датафрейме
    """
    pass





def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller
    Функция чтобы логотип отображался"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def select_folder_data():
    """
    Функция для выбора папки c данными
    :return:
    """
    global path_folder_data
    path_folder_data = filedialog.askdirectory()

def select_end_folder_alien():
    """
    Функция для выбора конечной папки куда будут складываться итоговые файлы
    :return:
    """
    global path_to_end_folder_alien
    path_to_end_folder_alien = filedialog.askdirectory()

def select_file_docx():
    """
    Функция для выбора файла Word
    :return: Путь к файлу шаблона
    """
    global file_docx
    file_docx = filedialog.askopenfilename(
        filetypes=(('Word files', '*.docx'), ('all files', '*.*')))

def select_file_data_alien():
    """
    Функция для выбора файла с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global file_data_alien
    # Получаем путь к файлу
    file_data_alien = filedialog.askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))

def select_files_data_xlsx():
    """
    Функция для выбора нескоьких файлов с данными на основе которых будет генерироваться документ
    :return: Путь к файлу с данными
    """
    global files_data_xlsx
    # Получаем путь файлы
    files_data_xlsx = filedialog.askopenfilenames(filetypes=(('Excel files', '*.xlsx'), ('all files', '*.*')))


def processing_report_alien():
    """
    Фугкция для получения данных по странам, количесту и сумме прошедших лечение
    :return:
    """
    try:
        _wb = openpyxl.load_workbook(file_data_alien, read_only=True) #загружаем чтобы узнать названия листов
        lst_sheets = _wb.sheetnames # получаем список листов
        _wb.close()

        country_df = pd.DataFrame(columns=range(10)) # создаем датафрейм в который будем добавлять данные
        all_rf_df  = pd.DataFrame(columns=range(10)) # создаем датафрейм в который будем добавлять данные по России
        for sheet_name in lst_sheets:
            df = pd.read_excel(file_data_alien, sheet_name=sheet_name, skiprows=1)
            if df.shape[1] !=9:
                raise WrongNumberColumn

            # Находим индекс первой пустой строки, если он есть,получаем список с значениями где есть пустые строки
            empty_row_index = np.where(df.isna().all(axis=1))
            if empty_row_index[0].tolist():
                row_index = empty_row_index[0][0]
                df = df.iloc[:row_index]

            df = df[~df['Страна'].isnull()]  # очищаем от не заполненных строк в колонке Страна
            df['Страна'] = df['Страна'].apply(lambda x: x.strip())

            # копируем все строки с РФ
            rf_df = df[df['Страна'] == 'РФ']
            rf_df.columns = range(9)
            rf_df[9] = sheet_name
            all_rf_df = pd.concat([all_rf_df,rf_df],axis=0)

            df = df[df['Страна'] != 'РФ']  # оставляем иностранцев
            df.columns = range(9)  # меняем названия чтобы избежать проблем с неправильным написанием
            df[9] = sheet_name # колонка с названием листа
            country_df = pd.concat([country_df, df], axis=0)  # добавляем в главный датафрейм

        #Обработка иностранцев
        itog_df = country_df.groupby(3).agg({3: 'count', 8: sum}) #  группируем

        itog_df.sort_values(by=8, ascending=False, inplace=True) # сортируем

        itog_df.columns = ['Количество контрагентов', 'Общая сумма'] # меняем названия колонок

        itog_df.index.name = 'Страна' # меняем название индекса

        itog_df = itog_df.reset_index()

        # Обработка Россия
        rf_itog_df = all_rf_df.groupby(3).agg({3: 'count', 8: sum}) #  группируем

        rf_itog_df.columns = ['Количество контрагентов', 'Общая сумма'] # меняем названия колонок

        rf_itog_df.index.name = 'Страна' # меняем название индекса

        rf_itog_df = rf_itog_df.reset_index()

        lst_country = country_df[3].unique() # список стран
        # генерируем текущее время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)

        wb = openpyxl.Workbook()
        main_sheet = wb['Sheet']
        main_sheet.title = 'Иностранцы'
        # Создаем заголовок для таблиц по странам
        country_columns = ['№','Номер документа','Вид операции','Страна','Контрагент','номер по прейскуранту','Наличные','Безналичные','Сумма документа','Лист']

        for r in dataframe_to_rows(itog_df, index=False, header=True):
            wb['Иностранцы'].append(r)
        wb['Иностранцы'].column_dimensions['A'].width = 30
        wb['Иностранцы'].column_dimensions['B'].width = 20
        wb['Иностранцы'].column_dimensions['C'].width = 20

        # создаем общий датафрейм
        all_df = pd.concat([itog_df,rf_itog_df],axis=0)
        all_df.sort_values(by=['Общая сумма'],ascending=False,inplace=True)

        wb.create_sheet(title='РФ + Иностранцы',index=1)
        for r in dataframe_to_rows(all_df, index=False, header=True):
            wb['РФ + Иностранцы'].append(r)
        wb['РФ + Иностранцы'].column_dimensions['A'].width = 30
        wb['РФ + Иностранцы'].column_dimensions['B'].width = 20
        wb['РФ + Иностранцы'].column_dimensions['C'].width = 20

        # создаем лист для РФ
        wb.create_sheet(title='Россия',index=2)
        all_rf_df.sort_values(by=8, ascending=False, inplace=True)  # сортируем
        all_rf_df.columns = country_columns
        for r in dataframe_to_rows(all_rf_df, index=False, header=True):
            wb['Россия'].append(r)
        wb['Россия'].column_dimensions['A'].width = 4
        wb['Россия'].column_dimensions['E'].width = 20
        wb['Россия'].column_dimensions['J'].width = 20

        # создаем листы
        for idx,country in enumerate(lst_country):
            wb.create_sheet(title=country, index=idx+3)

        # сохраняем страны на отдельные листы
        for country in lst_country:
            temp_df = country_df[country_df[3] == country] # получаем всех кто из этой страны
            temp_df.sort_values(by=8, ascending=False, inplace=True)  # сортируем
            temp_df.columns = country_columns
            for r in dataframe_to_rows(temp_df, index=False, header=True):
                wb[country].append(r)
            wb[country].column_dimensions['A'].width = 4
            wb[country].column_dimensions['E'].width = 20
            wb[country].column_dimensions['J'].width = 20

        wb.save(f'{path_to_end_folder_alien}/Отчет по иностранцам от {current_time}.xlsx')
        wb.close()
    except NameError:
        messagebox.showerror('Фебрис Обработка таблиц ver. 1.0',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
    except WrongNumberColumn:
        messagebox.showerror('Фебрис Обработка таблиц ver. 1.0',
                             f'Проверьте количество колонок на листе {sheet_name}. Должно быть 9 колонок !!!')
    except KeyError as e:
        messagebox.showerror('Фебрис Обработка таблиц ver. 1.0',
                             f'Не найдено значение {e.args}')
    except FileNotFoundError:
        messagebox.showerror('Фебрис Обработка таблиц ver. 1.0',
                             f'Перенесите файлы которые вы хотите обработать в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам')
    except PermissionError as e:
        messagebox.showerror('Фебрис Обработка таблиц ver. 1.0',
                             f'Закройте открытые файлы Excel {e.args}')
    else:
        messagebox.showinfo('Фебрис Обработка таблиц ver. 1.0',
                            'Данные успешно обработаны.')



if __name__ == '__main__':
    window = Tk()
    window.title('ЦОПП Бурятия Фебрис Обработка таблиц ver. 1.0')
    window.geometry('800x860')
    window.resizable(False, False)

    # Создаем объект вкладок

    tab_control = ttk.Notebook(window)

    # Создаем вкладку обработки данных для Приложения 6
    tab_alien_report = ttk.Frame(tab_control)
    tab_control.add(tab_alien_report, text='Отчет иностранцы')
    tab_control.pack(expand=1, fill='both')
    # Добавляем виджеты на вкладку Создание образовательных программ
    # Создаем метку для описания назначения программы
    lbl_hello = Label(tab_alien_report,
                      text='Центр опережающей профессиональной подготовки Республики Бурятия\n'
                           'Программа для подсчета количества иностранцев, суммы оказанны услуг по странам.')
    lbl_hello.grid(column=0, row=0, padx=10, pady=25)

    # Картинка
    path_to_img_alien = resource_path('logo.png')

    img_alien = PhotoImage(file=path_to_img_alien)
    Label(tab_alien_report,
          image=img_alien
          ).grid(column=1, row=0, padx=10, pady=25)

    # Создаем кнопку Выбрать файл с данными
    btn_choose_data_alien = Button(tab_alien_report, text='1) Выберите файл с данными', font=('Arial Bold', 20),
                             command=select_file_data_alien
                             )
    btn_choose_data_alien.grid(column=0, row=2, padx=10, pady=10)

    # Создаем кнопку для выбора папки куда будут генерироваться файлы

    btn_choose_end_folder_alien = Button(tab_alien_report, text='2) Выберите конечную папку', font=('Arial Bold', 20),
                                   command=select_end_folder_alien
                                   )
    btn_choose_end_folder_alien.grid(column=0, row=3, padx=10, pady=10)

    #Создаем кнопку обработки данных

    btn_proccessing_alien = Button(tab_alien_report, text='3) Обработать данные', font=('Arial Bold', 20),
                                  command=processing_report_alien
                                  )
    btn_proccessing_alien.grid(column=0, row=4, padx=10, pady=10)

    window.mainloop()